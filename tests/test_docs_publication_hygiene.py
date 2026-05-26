import unittest
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DOCS_DIR = PROJECT_ROOT / "docs"
REFERENCE_WORKBOOKS = [
    DOCS_DIR / "A5.2_Casos_Prueba_.xlsx",
    DOCS_DIR / "E5.2_Resultados_Validacion_Componentes.xlsx",
]
IGNORED_REFERENCE_PATTERNS = [
    "context/",
    "context\\",
    "deliverables/",
    "deliverables\\",
    "deployer.config",
    "validation/datasets/sources/",
    "validation/datasets/sources\\",
    "experiments/",
    "experiments\\",
    "artefactos de experiments",
    "logs.txt",
    "logs/",
    "runtime/",
    "runtime\\",
    "node_modules/",
    "playwright-report/",
    "test-results/",
    "blob-report/",
]
FORBIDDEN_SECRET_LITERALS = [
    "testing123",
    '"password"',
    "password fields",
    "password field",
    "secret key",
]
MALFORMED_SPANISH_PATTERNS = [
    "ciónes",
    "útiliz",
    "acciónes",
    "validaciónes",
    "integraciónes",
    "iteraciónes",
    "limitaciónes",
    "observaciónes",
]


class DocsPublicationHygieneTests(unittest.TestCase):
    def _workbook_strings(self):
        for workbook_path in REFERENCE_WORKBOOKS:
            workbook = load_workbook(workbook_path, data_only=False)
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str):
                            yield workbook_path, sheet.title, cell.coordinate, cell.value

    def test_reference_workbooks_do_not_point_to_ignored_local_paths(self):
        findings = []
        for workbook_path, sheet, coordinate, value in self._workbook_strings():
            normalized = value.lower()
            hits = [pattern for pattern in IGNORED_REFERENCE_PATTERNS if pattern.lower() in normalized]
            if hits:
                findings.append((workbook_path.name, sheet, coordinate, hits, value[:160]))

        self.assertEqual([], findings)

    def test_reference_workbooks_do_not_repeat_qa_actor(self):
        findings = []
        for workbook_path, sheet, coordinate, value in self._workbook_strings():
            parts = [part.strip().lower() for part in value.replace("\n", " / ").split("/") if part.strip()]
            seen = []
            for part in parts:
                if part in seen and part == "qa":
                    findings.append((workbook_path.name, sheet, coordinate, value[:160]))
                    break
                seen.append(part)

        self.assertEqual([], findings)

    def test_reference_workbooks_do_not_expose_literal_test_credentials(self):
        findings = []
        for workbook_path, sheet, coordinate, value in self._workbook_strings():
            normalized = value.lower()
            hits = [pattern for pattern in FORBIDDEN_SECRET_LITERALS if pattern.lower() in normalized]
            if hits:
                findings.append((workbook_path.name, sheet, coordinate, hits, value[:160]))

        self.assertEqual([], findings)

    def test_reference_workbooks_do_not_contain_malformed_spanish_after_normalization(self):
        findings = []
        for workbook_path, sheet, coordinate, value in self._workbook_strings():
            hits = [pattern for pattern in MALFORMED_SPANISH_PATTERNS if pattern in value]
            if hits:
                findings.append((workbook_path.name, sheet, coordinate, hits, value[:160]))

        self.assertEqual([], findings)

    def test_reference_workbook_sheet_names_are_polished(self):
        workbook = load_workbook(DOCS_DIR / "A5.2_Casos_Prueba_.xlsx", data_only=False)

        self.assertIn("A5.1_Funcionalidades_Ex.1", workbook.sheetnames)
        self.assertNotIn("A5.1_Funcionlidades_Ex.1", workbook.sheetnames)

    def test_a52_workbook_does_not_embed_media(self):
        with ZipFile(DOCS_DIR / "A5.2_Casos_Prueba_.xlsx") as workbook_zip:
            media_entries = [name for name in workbook_zip.namelist() if name.startswith("xl/media/")]

        self.assertEqual([], media_entries)

    def test_a52_functionality_matrix_uses_visual_checks(self):
        workbook = load_workbook(DOCS_DIR / "A5.2_Casos_Prueba_.xlsx", data_only=False)
        sheet = workbook["A5.1_Funcionalidades_Ex.1"]

        visual_checks = 0
        findings = []
        for row in sheet.iter_rows(min_row=2, min_col=4, max_col=6):
            for cell in row:
                if isinstance(cell.value, bool) or str(cell.value).upper() in {
                    "TRUE",
                    "FALSE",
                    "VERDADERO",
                    "FALSO",
                }:
                    findings.append((cell.coordinate, cell.value))
                if cell.value == "✓":
                    visual_checks += 1

        self.assertEqual([], findings)
        self.assertGreater(visual_checks, 0)

    def test_a52_component_case_sheets_share_reproduction_format(self):
        workbook = load_workbook(DOCS_DIR / "A5.2_Casos_Prueba_.xlsx", data_only=False)

        for sheet_name, component_title in [
            ("Ontology Hub", "Ontology Hub"),
            ("AI Model Hub", "AI Model Hub"),
            ("Semantic Virtualization", "Semantic Virtualization"),
        ]:
            sheet = workbook[sheet_name]
            self.assertEqual(
                ["ID", component_title, "Flujo", "Archivo de prueba (spec/runner)"],
                [sheet.cell(1, column).value for column in range(1, 5)],
            )

    def test_a52_workbook_keeps_inesdata_flows_in_one_sheet(self):
        workbook = load_workbook(DOCS_DIR / "A5.2_Casos_Prueba_.xlsx", data_only=False)

        self.assertIn("INESData", workbook.sheetnames)
        self.assertNotIn("Flujos_INESData", workbook.sheetnames)

        inesdata_sheet = workbook["INESData"]
        self.assertEqual(
            [
                "ID",
                "Nivel",
                "Caso / flujo",
                "Componente / ámbito",
                "Tipo de validación",
                "Actor / canal",
                "Precondiciones",
                "Procedimiento reproducible",
                "Resultado esperado",
                "Automatización / spec / runner",
                "Estado",
                "Evidencia",
                "Observaciones / trazabilidad",
            ],
            [inesdata_sheet.cell(1, column).value for column in range(1, 14)],
        )

        ids_by_level = {}
        for row in range(2, inesdata_sheet.max_row + 1):
            level = inesdata_sheet.cell(row, 2).value
            test_id = inesdata_sheet.cell(row, 1).value
            if level and test_id:
                ids_by_level.setdefault(level, set()).add(test_id)

        executable_ids = ids_by_level.get("Prueba ejecutable", set())
        flow_ids = ids_by_level.get("Flujo de integración", set())

        self.assertGreaterEqual(len(executable_ids), 18)
        self.assertGreaterEqual(len(flow_ids), 8)
        self.assertIn("DS-UI-AMH-OBS-02", executable_ids)
        self.assertIn("DS-UI-SV-01", executable_ids)
        self.assertIn("INT-OH-DS-03", executable_ids)
        self.assertIn("FLOW-AMH-DS-04", flow_ids)
        self.assertIn("FLOW-VS-DS-01", flow_ids)

    def test_reference_workbook_views_remain_scrollable(self):
        findings = []
        for workbook_path in REFERENCE_WORKBOOKS:
            workbook = load_workbook(workbook_path, data_only=False)
            for sheet in workbook.worksheets:
                hidden_rows = [
                    row
                    for row, dimension in sheet.row_dimensions.items()
                    if dimension.hidden
                ]
                hidden_columns = [
                    column
                    for column, dimension in sheet.column_dimensions.items()
                    if dimension.hidden
                ]
                freeze = sheet.freeze_panes
                freeze_row = 1
                if isinstance(freeze, str):
                    row_part = "".join(character for character in freeze if character.isdigit())
                    freeze_row = int(row_part) if row_part else 1
                if freeze_row > 6 or hidden_rows or hidden_columns:
                    findings.append(
                        (
                            workbook_path.name,
                            sheet.title,
                            freeze,
                            hidden_rows[:5],
                            hidden_columns[:5],
                        )
                    )

        self.assertEqual([], findings)

    def test_e52_summary_preserves_executive_layout(self):
        workbook = load_workbook(
            DOCS_DIR / "E5.2_Resultados_Validacion_Componentes.xlsx",
            data_only=False,
        )
        summary = workbook["00_Resumen"]

        self.assertEqual("Cierre por componente", summary["A21"].value)
        self.assertEqual("97/97 (100.0% cobertura técnica ejecutada)", summary["E28"].value)

    def test_e52_ontology_hub_sheets_are_merged_and_spanish_preconditions(self):
        workbook = load_workbook(
            DOCS_DIR / "E5.2_Resultados_Validacion_Componentes.xlsx",
            data_only=False,
        )

        self.assertIn("04_OntologyHub", workbook.sheetnames)
        self.assertNotIn("04_OntologyHub_PT5", workbook.sheetnames)
        self.assertNotIn("05_OntologyHub_UI", workbook.sheetnames)

        sheet = workbook["04_OntologyHub"]
        self.assertEqual(
            [
                "ID",
                "Nivel",
                "Caso / flujo",
                "Componente",
                "Tipo de validación",
                "Actor principal",
                "Precondiciones",
                "Procedimiento reproducible",
                "Resultado esperado",
                "Automatización / spec / runner",
                "Estado actual",
                "Evidencia disponible",
                "Fuente / entregable",
                "Observaciones",
            ],
            [sheet.cell(1, column).value for column in range(1, 15)],
        )

        levels = {}
        english_fragments = [
            " is ",
            " are ",
            "available",
            "deployed",
            "reachable",
            "exists",
            "at least",
            "system exposes",
        ]
        findings = []
        for row in range(2, sheet.max_row + 1):
            level = sheet.cell(row, 2).value
            if level:
                levels[level] = levels.get(level, 0) + 1
            preconditions = sheet.cell(row, 7).value
            if isinstance(preconditions, str):
                normalized = preconditions.lower()
                hits = [fragment for fragment in english_fragments if fragment in normalized]
                if hits:
                    findings.append((sheet.cell(row, 1).value, hits, preconditions))

        self.assertEqual(
            {
                "Soporte funcional": 2,
                "Caso PT5 oficial": 16,
                "Flujo UI funcional": 27,
            },
            levels,
        )
        self.assertEqual([], findings)

    def test_e52_ontology_hub_reflects_final_level6_without_historical_failures(self):
        workbook = load_workbook(
            DOCS_DIR / "E5.2_Resultados_Validacion_Componentes.xlsx",
            data_only=False,
        )
        sheet = workbook["04_OntologyHub"]
        headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
        status_column = headers.index("Estado actual") + 1
        observations_column = headers.index("Observaciones") + 1

        forbidden_observation_fragments = [
            "failed",
            "fallo",
            "fallando",
            "crash",
            "issue",
            "pendiente",
        ]
        findings = []
        statuses = {}
        for row in range(2, sheet.max_row + 1):
            test_id = sheet.cell(row, 1).value
            status = sheet.cell(row, status_column).value
            statuses[status] = statuses.get(status, 0) + 1
            if not str(status).startswith("Passed en Level 6 final"):
                findings.append((test_id, "status", status))

            observations = sheet.cell(row, observations_column).value
            if isinstance(observations, str):
                normalized = observations.lower()
                hits = [
                    fragment
                    for fragment in forbidden_observation_fragments
                    if fragment in normalized
                ]
                if hits:
                    findings.append((test_id, "observations", hits, observations[:160]))

        self.assertEqual([], findings)
        self.assertEqual(
            {"Passed en Level 6 final (experiment_2026-05-26_18-18-09)": 45},
            statuses,
        )

    def test_e52_sheet_numbering_is_continuous_without_color_legend(self):
        workbook = load_workbook(
            DOCS_DIR / "E5.2_Resultados_Validacion_Componentes.xlsx",
            data_only=False,
        )

        expected_sheets = [
            "00_Resumen",
            "01_Casos_Uso_A52",
            "02_Flujos_INESData",
            "03_INESData_UI",
            "04_OntologyHub",
            "05_AI_Model_Hub",
            "06_Virtualización",
            "07_Evidencias",
            "08_Trazabilidad_OH",
            "09_Estado_Cobertura",
            "10_Indice_Pruebas",
            "11_Level6_Results",
            "12_UNE_0087_Checklist",
        ]

        self.assertEqual(expected_sheets, workbook.sheetnames)
        self.assertNotIn("11_Leyenda_Colores", workbook.sheetnames)

    def test_markdown_docs_do_not_reference_context_directory(self):
        findings = []
        for markdown_path in DOCS_DIR.rglob("*.md"):
            content = markdown_path.read_text(encoding="utf-8")
            if "context/" in content:
                findings.append(markdown_path.relative_to(PROJECT_ROOT).as_posix())

        self.assertEqual([], findings)

    def test_markdown_docs_do_not_reference_old_workbook_sheet_typo(self):
        findings = []
        for markdown_path in DOCS_DIR.rglob("*.md"):
            content = markdown_path.read_text(encoding="utf-8")
            if "A5.1_Funcionlidades_Ex.1" in content:
                findings.append(markdown_path.relative_to(PROJECT_ROOT).as_posix())

        self.assertEqual([], findings)


if __name__ == "__main__":
    unittest.main()
